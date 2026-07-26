#!/usr/bin/env python3
from http.server import HTTPServer, BaseHTTPRequestHandler
import json, os, subprocess, tempfile, io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

BASE = "/Users/gajabmarketing/Library/CloudStorage/GoogleDrive-gajab@aeliyamarine.com/My Drive/Apps/Product-Video-Scraper/artifacts/api-server"
ENV_PATH = f"{BASE}/.env"
SCRAPER_SCRIPT = f"{BASE}/_scraper.py"
READ_EXCEL_SCRIPT = f"{BASE}/_read_excel.py"

try:
    for line in open(ENV_PATH).read().split("\n"):
        t = line.strip()
        if t and not t.startswith("#"):
            eq = t.find("=")
            if eq > 0: os.environ.setdefault(t[:eq], t[eq+1:])
except: pass

class Handler(BaseHTTPRequestHandler):
    def do_OPTIONS(self):
        self.send_response(204)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()

    def _cors(self, code=200, ctype="application/json"):
        self.send_response(code)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Content-Type", ctype)
        self.end_headers()

    def _json(self, data, code=200):
        self._cors(code)
        self.wfile.write(json.dumps(data).encode())

    def _read_body(self):
        length = int(self.headers.get("Content-Length", 0))
        return self.rfile.read(length) if length else b""

    def _parse_multipart(self, body, boundary):
        boundary = boundary.encode() if isinstance(boundary, str) else boundary
        b = b"--" + boundary
        parts = body.split(b)[1:]
        for part in parts:
            if part.startswith(b"--"): break
            idx = part.find(b"\r\n\r\n")
            if idx == -1: continue
            headers = part[:idx].decode("utf-8", errors="ignore")
            content = part[idx+4:]
            if content.endswith(b"\r\n"): content = content[:-2]
            if "name=\"file\"" in headers.replace(" ", "") or 'name=file' in headers:
                return content
        return None

    def do_POST(self):
        path = self.path.split("?")[0]
        content_type = self.headers.get("Content-Type", "")
        
        if path in ("/scraper/flipkart/upload", "/api/scraper/flipkart/upload"):
            if "multipart/form-data" not in content_type:
                self._json({"error": "Expected multipart/form-data"}, 400)
                return
            boundary = content_type.split("boundary=")[1].strip()
            raw = self._read_body()
            filedata = self._parse_multipart(raw, boundary)
            if not filedata:
                self._json({"error": "No file found"}, 400)
                return
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
                f.write(filedata)
                tmp = f.name
            try:
                r = subprocess.run(["python3", READ_EXCEL_SCRIPT, tmp], capture_output=True, text=True, timeout=30)
                os.unlink(tmp)
                data = json.loads(r.stdout)
                if data.get("error"): self._json({"error": data["error"]}, 500)
                else: self._json(data)
            except Exception as e:
                self._json({"error": str(e)}, 500)
            return

        if path in ("/scraper/scrape", "/api/scraper/scrape", "/scraper/flipkart/scrape", "/api/scraper/flipkart/scrape"):
            try:
                data = json.loads(self._read_body())
                urls = data.get("urls", [])
                if not urls:
                    self._json({"error": "URLs required"}, 400)
                    return
                results = []
                for i in range(0, len(urls), 2):
                    for url in urls[i:i+2]:
                        try:
                            r = subprocess.run(["python3", SCRAPER_SCRIPT, url], capture_output=True, text=True, timeout=180, env={**os.environ})
                            result = json.loads(r.stdout)
                            if not result.get("url"): result["url"] = url
                            results.append(result)
                        except Exception as e:
                            results.append({"status": "failed", "error": str(e), "url": url})
                self._json({"products": results})
            except Exception as e:
                self._json({"error": str(e)}, 500)
            return

        if path in ("/scraper/flipkart/export", "/api/scraper/flipkart/export"):
            try:
                data = json.loads(self._read_body())
                products = data.get("products", [])
                max_imgs = max((len(p.get("images") or []) for p in products), default=0)
                if not max_imgs:
                    for p in products:
                        img = p.get("imageUrl") or ""
                        if img: max_imgs = max(max_imgs, 1)
                wb = Workbook()
                ws = wb.active
                ws.title = "Products"
                headers = ["ID","Title","Description","Price","HSN","GST","Dimensions","Weight","Specifications","Variants"]
                for i in range(max_imgs):
                    headers.append(f"Image {i+1}")
                headers.append("Product URL")
                for col, h in enumerate(headers, 1):
                    c = ws.cell(row=1, column=col, value=h)
                    c.font = Font(bold=True, color="FFFFFF")
                    c.fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
                for ri, p in enumerate(products, 2):
                    specs = p.get("specifications") or {}
                    specs_s = "\n".join(f"{k}: {v}" for k,v in specs.items()) if isinstance(specs, dict) else str(specs)
                    imgs = p.get("images") or []
                    if not imgs and p.get("imageUrl"):
                        imgs = [p.get("imageUrl")]
                    vals = [p.get("id",""), p.get("title",""), p.get("description",""), p.get("price",""), p.get("hsn",""), p.get("gst",""), p.get("dimensions",""), p.get("weight",""), specs_s, p.get("variants","")]
                    for i in range(max_imgs):
                        vals.append(imgs[i] if i < len(imgs) else "")
                    vals.append(p.get("url",""))
                    for col, v in enumerate(vals, 1):
                        ws.cell(row=ri, column=col, value=str(v) if v else "")
                buf = io.BytesIO()
                wb.save(buf)
                data = buf.getvalue()
                self.send_response(200)
                self.send_header("Access-Control-Allow-Origin", "*")
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", 'attachment; filename="flipkart-products.xlsx"')
                self.send_header("Content-Length", str(len(data)))
                self.end_headers()
                self.wfile.write(data)
            except Exception as e:
                self._json({"error": str(e)}, 500)
            return

        self._json({"error": "Not found"}, 404)

    def do_GET(self):
        self._json({"error": "Not found"}, 404)

HTTPServer(("0.0.0.0", 8080), Handler).serve_forever()

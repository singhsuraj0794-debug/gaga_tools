#!/usr/bin/env python3
import json, sys
from openpyxl import load_workbook

filepath = sys.argv[1]
wb = load_workbook(filepath, read_only=True, data_only=True)
ws = wb.active
urls = []
url_cols = set()

for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
    if row_idx == 1:
        for col_idx, cell in enumerate(row, 1):
            val = str(cell or "").strip().lower()
            if any(k in val for k in ["url", "link", "product", "flipkart", "meesho"]):
                url_cols.add(col_idx)
        if not url_cols:
            url_cols = set(range(1, len(row) + 1))
        continue
    for col_idx in url_cols:
        if col_idx > len(row): continue
        val = str(row[col_idx - 1] or "").strip()
        if val and ("http" in val or "flipkart.com" in val or "meesho.com" in val):
            if not val.startswith("http"): val = "https://" + val
            urls.append(val)

if not urls:
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            val = str(cell or "").strip()
            if val and "http" in val: urls.append(val)

unique = list(dict.fromkeys(urls))
valid = [u for u in unique if "flipkart.com" in u or "meesho.com" in u]
print(json.dumps({"urls": unique, "totalUrls": len(unique), "validUrls": len(valid)}))

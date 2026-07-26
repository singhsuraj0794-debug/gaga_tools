from __future__ import annotations

import logging
import os
from typing import Any

import config

logger = logging.getLogger(__name__)


def read_sheet(
    sheet_id: str | None = None,
    sheet_name: str | None = None,
    file_path: str | None = None,
) -> tuple[list[list[str]], Any, str]:
    if file_path:
        return _read_local(file_path)
    if sheet_id:
        return _read_google(sheet_id, sheet_name)
    raise ValueError("Provide either --sheet-id or --file")


def _read_google(
    sheet_id: str,
    sheet_name: str | None,
) -> tuple[list[list[str]], Any, str]:
    import gspread
    from google.oauth2.service_account import Credentials

    creds_path = config.GOOGLE_SHEET_CREDENTIALS
    if not os.path.exists(creds_path):
        raise FileNotFoundError(
            f"Credentials file not found at {creds_path}. "
            "Set GOOGLE_SHEET_CREDENTIALS env or place credentials.json in the project root."
        )

    scope = [
        "https://spreadsheets.google.com/feeds",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = Credentials.from_service_account_file(creds_path, scopes=scope)
    client = gspread.authorize(creds)
    sh = client.open_by_key(sheet_id)
    ws = sh.sheet1 if sheet_name is None else sh.worksheet(sheet_name)
    rows = ws.get_all_values()
    logger.info(
        "Read %s rows (including header) from Google Sheet '%s'",
        len(rows),
        sh.title,
    )
    return rows, ws, "google"


def _read_local(file_path: str) -> tuple[list[list[str]], Any, str]:
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".csv":
        import pandas as pd

        df = pd.read_csv(file_path, dtype=str)
        rows = [list(df.columns)] + df.astype(str).values.tolist()
        wb = file_path
        logger.info("Read %s rows from CSV", len(rows))
        return rows, wb, "csv"

    import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    rows = [[(cell.value or "") for cell in row] for row in ws.iter_rows()]
    logger.info("Read %s rows (including header) from Excel file", len(rows))
    return rows, wb, "excel"


def find_url_column(headers: list[str]) -> int | None:
    url_keywords = ["url", "link", "product url", "product link", "page url"]
    for i, h in enumerate(headers):
        h_clean = h.strip().lower()
        if h_clean in url_keywords:
            return i
    for i, h in enumerate(headers):
        if "url" in h_clean:
            return i
    return None


def write_results(
    ws: Any,
    rows: list[list[str]],
    source_type: str,
    url_col: int,
    results: list[dict[str, str] | None],
    header_row: int = 0,
) -> None:
    if source_type == "google":
        _write_google(ws, rows, url_col, results)
    elif source_type == "excel":
        _write_excel(ws, rows, url_col, results)
    elif source_type == "csv":
        _write_csv(ws, rows, url_col, results)


def _column_letter(index: int) -> str:
    letters = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    if index < 26:
        return letters[index]
    return letters[index // 26 - 1] + letters[index % 26]


def _write_google(
    ws: Any,
    rows: list[list[str]],
    url_col: int,
    results: list[dict[str, str] | None],
) -> None:
    headers = list(rows[0]) if rows else []

    for col_name in config.RESULT_COLUMNS:
        found = any(h.strip().lower() == col_name.lower() for h in headers)
        if not found:
            headers.append(col_name)

    grid = [headers]
    for row_idx, row in enumerate(rows[1:]):
        extended = list(row)
        while len(extended) < len(headers):
            extended.append("")
        result = results[row_idx] if row_idx < len(results) else None
        if result is not None:
            for col_name in config.RESULT_COLUMNS:
                col_idx = next(
                    i for i, h in enumerate(headers)
                    if h.strip().lower() == col_name.lower()
                )
                key = _result_key(col_name)
                extended[col_idx] = result.get(key, "") or ""
        grid.append(extended)

    range_end = _column_letter(len(headers) - 1)
    ws.update(f"A1:{range_end}{len(grid)}", grid, value_input_option="USER_ENTERED")


def _write_excel(
    wb: Any,
    rows: list[list[str]],
    url_col: int,
    results: list[dict[str, str] | None],
) -> None:
    ws = wb.active
    headers = rows[0] if rows else []
    existing = set(h.lower().strip() for h in headers)

    next_col = len(headers)
    col_map: dict[str, int] = {}

    for col_name in config.RESULT_COLUMNS:
        found = False
        for i, h in enumerate(headers):
            if h.strip().lower() == col_name.lower():
                col_map[col_name] = i
                found = True
                break
        if not found:
            col_map[col_name] = next_col
            ws.cell(row=1, column=next_col + 1, value=col_name)
            next_col += 1

    for row_idx, result in enumerate(results):
        if result is None:
            continue
        excel_row = row_idx + 2
        for col_name in config.RESULT_COLUMNS:
            col_idx = col_map[col_name]
            key = _result_key(col_name)
            val = result.get(key, "")
            ws.cell(row=excel_row, column=col_idx + 1, value=val or "")

    wb.save(wb.path if hasattr(wb, "path") else wb)


def _write_csv(
    path: str,
    rows: list[list[str]],
    url_col: int,
    results: list[dict[str, str] | None],
) -> None:
    import pandas as pd

    df = pd.DataFrame(rows[1:], columns=rows[0])
    for i, result in enumerate(results):
        if result is None:
            continue
        for col_name in config.RESULT_COLUMNS:
            key = _result_key(col_name)
            df.at[i, col_name] = result.get(key, "")

    df.to_csv(path, index=False)
    logger.info("Results written to %s", path)


def _result_key(col_name: str) -> str:
    mapping = {
        "Title": "title",
        "Description": "description",
        "Image URL": "image_url",
        "Dimensions": "dimensions",
        "HSN": "hsn",
        "Status": "status",
        "Error": "error",
    }
    return mapping.get(col_name, col_name.lower())
